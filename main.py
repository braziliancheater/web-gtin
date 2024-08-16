from flask import Flask, request, jsonify, render_template, make_response, send_file, url_for
from flask_sqlalchemy import SQLAlchemy
import requests
from requests.adapters import HTTPAdapter
from requests.exceptions import SSLError
import xml.etree.ElementTree as ET
from requests.packages.urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os
import time

# Desabilita os warnings de certificado 
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///banco_local.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# ===========
#   Tabelas
# ===========
class dados_ean13(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    empresa = db.Column(db.String(2), nullable=False)
    item = db.Column(db.String(15), nullable=False)
    den_item = db.Column(db.String(120), nullable=False)
    tip_cod_barra = db.Column(db.String(5), nullable=False)
    cod_barra = db.Column(db.String(20), nullable=True)
    cod_barra_trib = db.Column(db.String(20), nullable=True)
    cod_gtin = db.Column(db.Integer, nullable=True)

    def __repr__(self):
        return f'<dados_ean13 {self.id}>'

class dados_dun14(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    empresa = db.Column(db.String(2), nullable=False)
    item = db.Column(db.String(15), nullable=False)
    den_item = db.Column(db.String(120), nullable=False)
    tip_cod_barra = db.Column(db.String(5), nullable=False)
    cod_barra = db.Column(db.String(20), nullable=True)
    cod_barra_trib = db.Column(db.String(20), nullable=True)
    cod_gtin = db.Column(db.Integer, nullable=True)

    def __repr__(self):
        return f'<dados_dun14 {self.id}>'

class gtin_dados(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    gtin = db.Column(db.String(20), nullable=False)
    status = db.Column(db.String(250), nullable=False)
    ncm = db.Column(db.String(8), nullable=True)
    cest = db.Column(db.String(7), nullable=True)
    data = db.Column(db.DateTime, nullable=False, default=db.func.current_timestamp())

    def __repr__(self):
        return f'<gtin_dados {self.id}>'

class logs(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.DateTime, nullable=False, default=db.func.current_timestamp())
    mensagem = db.Column(db.String(250), nullable=False)
    ip = db.Column(db.String(15), nullable=False)

    def __repr__(self):
        return f'<logs {self.id}>'

class propriedades(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(50), nullable=False)
    valor = db.Column(db.String(50), nullable=False)

    def __repr__(self):
        return f'<propriedades {self.id}>'

# =============
#  Tabelas Itens
# =============
# Atualizar tabelas
@app.route('/atualizar_tabelas')
def atualizar_tabelas():
    try:
        log("fazendo requisições para popular a tabela...")
        try:
            log("requisição do Dun14")
            dun14_url = "http://web-app:8080/wsriclan/portal_dun14.rule?sys=WEB&fonte=site"
            resposta = requests.get(dun14_url)
            resposta.raise_for_status() # da erro caso status != 200
            dados_dun14_data = resposta.json()

            for item in dados_dun14_data:
                # verifica se o item já existe no banco de dados
                cod_barra = item['cod_barra'].strip() if item['cod_barra'] else None
                cod_barra_trib = item['cod_barra_trib'].strip() if item['cod_barra_trib'] else None

                # Filter based on available cod_barra and cod_barra_trib
                verifica_item = None
                if cod_barra:
                    verifica_item = dados_dun14.query.filter_by(cod_barra=cod_barra).first()
                if not verifica_item and cod_barra_trib:
                    verifica_item = dados_dun14.query.filter_by(cod_barra_trib=cod_barra_trib).first()

                if verifica_item:
                    log(f"{cod_barra or cod_barra_trib} já existe no banco de dados")
                    continue

                linha = dados_dun14(
                    empresa=item['empresa'],
                    item=item['item'].strip() if item['item'] else None,
                    den_item=item['den_item'].strip() if item['den_item'] else None,
                    tip_cod_barra=item['tip_cod_barra'],
                    cod_barra=cod_barra,
                    cod_barra_trib=cod_barra_trib
                )
                db.session.add(linha)
            db.session.commit()
        except Exception as e:
            log(f"erro ao fazer requisição do Dun14, {e}")
            db.session.rollback()
        
        try:
            log("requisição do Ean13")
            ean13 = "http://web-app:8080/wsriclan/portal_ean13.rule?sys=WEB&fonte=site"
            resposta = requests.get(ean13)
            resposta.raise_for_status() # da erro caso status != 200
            dados = resposta.json()

            for item in dados:
                cod_barra = item['cod_barra'].strip() if item['cod_barra'] else None
                cod_barra_trib = item['cod_barra_trib'].strip() if item['cod_barra_trib'] else None

                # Filter based on available cod_barra and cod_barra_trib
                verifica_item = None
                if cod_barra:
                    verifica_item = dados_ean13.query.filter_by(cod_barra=cod_barra).first()
                if not verifica_item and cod_barra_trib:
                    verifica_item = dados_ean13.query.filter_by(cod_barra_trib=cod_barra_trib).first()

                if verifica_item:
                    log(f"{cod_barra or cod_barra_trib} já existe no banco de dados")
                    continue
            
                linha = dados_ean13(
                    empresa=item['empresa'],
                    item=item['item'].strip(),
                    den_item=item['den_item'].strip(),
                    tip_cod_barra=item['tip_cod_barra'],
                    cod_barra=cod_barra,
                    cod_barra_trib=cod_barra_trib
                )
                db.session.add(linha)
            db.session.commit()
        except:
            db.session.rollback()

        return jsonify({'mensagem': 'tabelas atualizadas com sucesso'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensagem': str(e)}), 500

@app.route('/limpar_tabelas')
def limpar_tabelas():
    try:
        db.session.query(dados_dun14).delete()
        db.session.query(dados_ean13).delete()
        db.session.commit()
        return jsonify({'mensagem': 'tabelas limpas com sucesso'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensagem': str(e)}), 500

# ===========
# Front End
# ===========

@app.route('/')
def home():
    criar_log("Acesso a página inicial")
    return render_template('index.html')

@app.route('/admin')
def admin():
    criar_log("Acesso a página de administração")
    return render_template('admin.html')

# Home/Listagem
@app.route('/listagem')
def listagem():
    criar_log("Acesso a página de listagem")

    ean13_data = db.session.query(dados_ean13, gtin_dados).outerjoin(gtin_dados, dados_ean13.cod_gtin == gtin_dados.id).all()
    dun14_data = db.session.query(dados_dun14, gtin_dados).outerjoin(gtin_dados, dados_dun14.cod_gtin == gtin_dados.id).all()

    mensagem = '''
                <div  class="alert alert-danger" role="alert">
                <h4 class="alert-heading">
                    <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" class="lucide lucide-badge-alert"><path d="M3.85 8.62a4 4 0 0 1 4.78-4.77 4 4 0 0 1 6.74 0 4 4 0 0 1 4.78 4.78 4 4 0 0 1 0 6.74 4 4 0 0 1-4.77 4.78 4 4 0 0 1-6.75 0 4 4 0 0 1-4.78-4.77 4 4 0 0 1 0-6.76Z"/><line x1="12" x2="12" y1="8" y2="12"/><line x1="12" x2="12.01" y1="16" y2="16"/></svg>
                    Aviso
                </h4>
                <a>Os dados abaixo ainda não foram confrontados com o GTIN.</a>
            </div>
    '''

    # mostra mensagem de alerta caso não existam GTINs
    gtin = gtin_dados.query.all()
    if len(gtin) == 0:
        return render_template('listagem.html', ean13_data=ean13_data, dun14_data=dun14_data, mensagem_gtin=mensagem)
    else:
        return render_template('listagem.html', ean13_data=ean13_data, dun14_data=dun14_data, mensagem_gtin=None)

# ===========
#  Gtin
# ===========

@app.route('/apagar_gtin')
def apagar_gtin():
    try:
        db.session.query(gtin_dados).delete()
        db.session.commit()
        return jsonify({'mensagem': 'GTINs apagados com sucesso'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensagem': str(e)}), 500

@app.route('/atualizar_gtin')
def atualizar_gtin():
    url = 'https://dfe-servico.svrs.rs.gov.br/ws/ccgConsGTIN/ccgConsGTIN.asmx?wsdl'
    xml_request_template = """
<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
                  xmlns:xsd="http://www.w3.org/2001/XMLSchema"
                  xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">
  <soap12:Body>
    <ccgConsGTIN xmlns="http://www.portalfiscal.inf.br/nfe/wsdl/ccgConsGtin">
      <nfeDadosMsg>
        <consGTIN versao="1.00" xmlns="http://www.portalfiscal.inf.br/nfe"><GTIN>{gtin}</GTIN></consGTIN>
      </nfeDadosMsg>
    </ccgConsGTIN>
  </soap12:Body>
</soap12:Envelope>
"""
    # configuração da sessão HTTPs
    session = requests.Session()
    retries = Retry(total=1, backoff_factor=1, status_forcelist=[502, 503, 504])
    adapter = HTTPAdapter(max_retries=retries)
    session.mount('http://', adapter)
    session.mount('https://', adapter)

    def processar_dados(dados):
        tempo = float(propriedades.query.filter_by(nome="velocidade").first().valor)
        for item in dados:
            code = item.cod_barra or item.cod_barra_trib
            if code:
                verifica_gtin = gtin_dados.query.filter_by(gtin=code).first()
                if verifica_gtin:
                    log(f"GTIN {code} já existe no banco de dados")
                    item.cod_gtin = verifica_gtin.id
                    db.session.commit()
                    continue

                xml_request = xml_request_template.format(gtin=code)
                try:
                    caminhos = obter_certificados()
                    response = session.post(url, cert=(caminhos[0], caminhos[1]), verify=False, data=xml_request, headers={'Content-Type': 'text/xml'})
                    response.raise_for_status()

                    # parse a resposta em XML
                    root = ET.fromstring(response.content)

                    # inicializar status e ncm
                    status = None
                    ncm = None
                    cest = None

                    # encontrar os elementos desejados no XML e imprimir na tela
                    for elem in root.iter():
                        if elem.tag.endswith('cStat'):
                            if elem.text == '9490':
                                status = elem.text + " - Produto existe"
                            elif elem.text == '9491':
                                status = elem.text + " - GTIN com dígito verificador inválido"
                                log(f"Status: {elem.text} - GTIN com dígito verificador inválido")
                            elif elem.text == '9492':
                                status = elem.text + " - GTIN não possui prefixo 789 ou 790 (Brasil)"
                                log(f"Status: {elem.text} - GTIN não possui prefixo 789 ou 790 (Brasil)")
                            elif elem.text == '9493':
                                status = elem.text + " - CNPJ/CPF do Certificado de Transmissão não é emitente de NF-e ou NFC-e"
                                log(f"Status: {elem.text} - CNPJ/CPF do Certificado de Transmissão não é emitente de NF-e ou NFC-e")
                            elif elem.text == '9494':
                                status = elem.text + " - GTIN inexistente no Cadastro Centralizado de GTIN (CCG)"
                                log(f"Status: {elem.text} - GTIN inexistente no Cadastro Centralizado de GTIN (CCG)")
                            elif elem.text == '9495':
                                #status = elem.text + " - GTIN existe no CCG com situação inválida. Solicitar ao Dono da Marca que entre em contato com a GS1 (Inativo)"
                                status = elem.text + " - GTIN existe no CCG com situação inválida."
                                log(f"Status: {elem.text} - GTIN existe no CCG com situação inválida. Solicitar ao Dono da Marca que entre em contato com a GS1")
                            elif elem.text == '9496':
                                #status = elem.text + " - GTIN existe no CCG, mas dono da marca não autorizou a publicação das informações. Entrar em contato com o Dono da Marca"
                                status = elem.text + " - GTIN existe no CCG, mas dono da marca não autorizou a publicação das informações."
                                log(f"Status: {elem.text} - GTIN existe no CCG, mas dono da marca não autorizou a publicação das informações. Entrar em contato com o Dono da Marca")
                            elif elem.text == '9497':
                                status = elem.text + " - GTIN existe no CCG com NCM não informado"
                                log(f"Status: {elem.text} - GTIN existe no CCG com NCM não informado")
                            elif elem.text == '9498':
                                status = elem.text + " - GTIN existe no CCG com NCM inválido"
                                log(f"Status: {elem.text} - GTIN existe no CCG com NCM inválido")
                        elif elem.tag.endswith('xProd'):
                            produto = elem.text
                            log(f"Produto: {elem.text}")
                        elif elem.tag.endswith('NCM'):
                            ncm = elem.text
                            log(f"NCM: {elem.text}")
                        if elem.tag.endswith('CEST'):
                            cest = elem.text
                            log(f"CEST: {elem.text}")
                        elif elem.tag.endswith('Falha'):
                            log("Erro falha XML")

                    # verificar se o status foi definido
                    if status is not None:
                        gtin_entry = gtin_dados(
                            gtin=code,
                            status=status,
                            ncm=ncm
                        )
                        db.session.add(gtin_entry)
                        db.session.commit()
                        item.cod_gtin = gtin_entry.id
                        db.session.commit()

                    # tempo entre as consultas
                    time.sleep(tempo)

                except (SSLError, requests.exceptions.RequestException) as e:
                    log(f"Erro SSL: {e}")
                    continue  # Continue instead of exit to process the remaining items

    try:
        # processar dados_dun14
        log("Processando dados_dun14...")
        dun14_data = dados_dun14.query.all()
        processar_dados(dun14_data)

        # processar dados_ean13
        log("Processando dados_ean13...")
        ean13_data = dados_ean13.query.all()
        processar_dados(ean13_data)

        return jsonify({'mensagem': 'GTINs atualizados com sucesso'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensagem': str(e)}), 500

# ==========
#  Testes Q/A
# ==========

@app.route('/testar_gtin')
def gtin_teste():
    gtin_r = request.args.get('gtin')
    url = 'https://dfe-servico.svrs.rs.gov.br/ws/ccgConsGTIN/ccgConsGTIN.asmx?wsdl'
    xml_request_template = """
<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
                  xmlns:xsd="http://www.w3.org/2001/XMLSchema"
                  xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">
  <soap12:Body>
    <ccgConsGTIN xmlns="http://www.portalfiscal.inf.br/nfe/wsdl/ccgConsGtin">
      <nfeDadosMsg>
        <consGTIN versao="1.00" xmlns="http://www.portalfiscal.inf.br/nfe"><GTIN>{gtin}</GTIN></consGTIN>
      </nfeDadosMsg>
    </ccgConsGTIN>
  </soap12:Body>
</soap12:Envelope>
"""
    # configuração da sessão HTTP
    session = requests.Session()
    retries = Retry(total=1, backoff_factor=1, status_forcelist=[502, 503, 504])
    adapter = HTTPAdapter(max_retries=retries)
    session.mount('http://', adapter)
    session.mount('https://', adapter)

    xml_request = xml_request_template.format(gtin=gtin_r)
    try:
        caminhos = obter_certificados()
        response = session.post(url, cert=(caminhos[0], caminhos[1]), verify=False, data=xml_request, headers={'Content-Type': 'text/xml'})
        response.raise_for_status()

        # verificar se o tipo de resposta é XML
        if 'xml' not in response.headers.get('Content-Type', ''):
            return f"Tipo invalido na resposta: {response.headers.get('Content-Type')}", 500

        # converte a resposta em texto
        response_content = response.content.decode('utf-8')

        # parse a resposta em XML
        return make_response((response_content, 200, {'Content-Type': 'text/xml'}))
    except (SSLError, requests.exceptions.RequestException) as e:
        return str(e), 500

# Tela de testes
@app.route('/testes', methods=['GET'])
def testes():        
    urls = ['http://web-app:8080/wsriclan/portal_ean13.rule?sys=WEB&fonte=site', 'http://web-app:8080/wsriclan/portal_dun14.rule?sys=WEB&fonte=site']
    status = []

    for url in urls:
        requisicao = requests.get(url).status_code
        status.append(requisicao)

    resposta = {
        "ean13": status[0],
        "dun14": status[1]
        }
    return resposta, 200

# =================
#  Logs de Acesso
# =================

def criar_log(mensagem):
    try:
        if mensagem is None:
            return False
        ip = request.remote_addr
        log = logs(mensagem=mensagem, ip=ip)
        db.session.add(log)
        db.session.commit()
        return True
    except Exception as e:
        db.session.rollback()
        return False
    
# ===============
# Exportar Excel
# ===============

@app.route('/exportar_excel')
def exportar_excel():
    import pandas as pd
    try:
        # obtem todos os dados
        ean13_data = dados_ean13.query.all()
        dun14_data = dados_dun14.query.all()

        # converte as tabelas em dados do pandas
        ean13_df = pd.DataFrame([{
            'empresa': d.empresa, 
            'item': d.item, 
            'den_item': d.den_item, 
            'tip_cod_barra': d.tip_cod_barra,
            'cod_barra': d.cod_barra,
            'cod_barra_trib': d.cod_barra_trib,
            'cod_gtin': d.cod_gtin
        } for d in ean13_data])

        dun14_df = pd.DataFrame([{
            'empresa': d.empresa, 
            'item': d.item, 
            'den_item': d.den_item, 
            'tip_cod_barra': d.tip_cod_barra,
            'cod_barra': d.cod_barra,
            'cod_barra_trib': d.cod_barra_trib,
            'cod_gtin': d.cod_gtin
        } for d in dun14_data])

        index_df = pd.DataFrame([{
            '': ''
        }])

        # salvar em excel
        caminho_pasta = 'arquivos_temporarios'
        if not os.path.exists(caminho_pasta):
            os.makedirs(caminho_pasta)

        caminho_arquivo = os.path.join(caminho_pasta, "dados_gtin.xlsx")

        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            index_df.to_excel(writer, sheet_name='Inicio', index=False)
            ean13_df.to_excel(writer, sheet_name='EAN13', index=False)
            dun14_df.to_excel(writer, sheet_name='DUN14', index=False)

            # adiciona um quadrado em volta do texto
            workbook = writer.book
            primeira_pagina = workbook['Inicio']
            primeira_pagina.insert_rows(0)

            primeira_pagina['A1'] = "Os dados desta planilha não são atualizados automaticamente, utilize o site para dados atualizados."
            primeira_pagina['A1'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            primeira_pagina['A1'].font = Font(color="FFFFFF")
            primeira_pagina['A2'] = "Dados obtidos no dia " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) 
            primeira_pagina['A2'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            primeira_pagina['A2'].font = Font(color="FFFFFF")
            primeira_pagina['A3'] = "GTIN Web - Gabriel" 
            primeira_pagina['A3'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            primeira_pagina['A3'].font = Font(color="FFFFFF")

            primeira_pagina.column_dimensions['A'].width = 120
            # centraliza o texto
            primeira_pagina['A1'].alignment = primeira_pagina['A2'].alignment = primeira_pagina['A3'].alignment = Alignment(horizontal='center', vertical='center')

        # manda arquivo para o usuário baixar
        response = send_file(caminho_arquivo, as_attachment=True, download_name='dados_gtin.xlsx')
        criar_log("Usuario exportou dados para Excel")

        # limpa o arquivo da pasta para, não termos arquivos antigos
        @response.call_on_close
        def cleanup():
            try:
                if os.path.exists(caminho_arquivo):
                    os.remove(caminho_arquivo)
            except Exception as e:
                log(f"Erro ao deletar o arquivo: {e}")

        # manda de volta para tela de admin
        response.headers['Location'] = url_for('admin')

        return response

    except Exception as e:
        log(f"Erro ao exportar para Excel: {e}")
        return jsonify({'mensagem': str(e)}), 500
    
# =======
#  Utils
# =======
def testar_certificados():
    log("testando se certificados existem...")
    caminhos = ["cert", "cert/cert.pem", "cert/key.pem"]
    
    try:
        for caminho in caminhos:
            teste = os.path.exists(caminho)
            if teste is False:
                log(f"[ERRO] Caminho {caminho} não existe...")
                return False
            else:
                continue
        return True
    except Exception as e:
        log(f"[ERRO] {e}")
        return False

def obter_certificados():
    log("obtendo certificados...")
    caminhos = ["cert/cert.pem", "cert/key.pem"]

    if (testar_certificados() is False):
        log("certificados não existem, não é possível continuar")
        exit()
    else:
        log("certificados existem, continuando...")

    return caminhos

def verifica_propriedades():
    try:
        propriedade = propriedades.query.all()
        if len(propriedade) == 0:
            log("Propriedades não existem, criando...")
            velocidade = propriedades(nome="velocidade", valor="0.1")
            db.session.add(velocidade)
            db.session.commit()
        else:
            log("Propriedades existem, continuando...")
    except Exception as e:
        log(f"Erro ao verificar propriedades, {e}")

def log(message):
    vermelho_negrito = "\x1b[31;1m"
    branco = "\x1b[37m"
    reset = "\x1b[0m"
    # hora atual
    hora = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    print(f"{vermelho_negrito}[{hora}][web-gtin] {branco}{message}{reset}")

def logo():
    import os 
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')
    print("""
               xxxx
              xx   
 Gabriel     x    
          xxxxx    
          x       Todos   
         xx     os Direitos
        xx     Reservados
   Riclan S.A      
       xx          
      x            
  xxxx     Web-GTIN        
 xx                
 x                 
x                  
""")

# na primeira requisição feita pelo site, verifica se as tabelas estão criadas
with app.app_context():
    logo()
    log("verificando se as tabelas existem...")
    db.create_all()
    log("tabelas criadas, continuando...")

    # verificando se os certificados existem
    if (testar_certificados() is False):
        log("certificados não existem, não é possível continuar")
        exit()
    else:
        log("certificados existem, continuando...")

    # verificando se as propriedades existem
    verifica_propriedades()

if __name__ == '__main__':
    app.run(debug=True)